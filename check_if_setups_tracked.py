#! python3

import pyodbc
import openpyxl
import send_email
from datetime import *

joe = 'jbriggs@catiglass.com;'
lazarrick = 'lstokes@catiglass.com;'
penny = 'penny.quandt@catiglass.com;'
slawek = 'slawek@catiglass.com;'
send_to = joe + lazarrick + penny + slawek
subject = 'Missed Setup Tracking jobs'
default_msg = "No missed tracking found today."
found_msg = "Today we found the following missed setups:\n"

trackedlist = []
nontrackedlist = []
#location of Setup Tracking spreadsheet
setupPath = "G:\\3 - Production Departments\\11- Scheduling\\Grinding\\Setup_Tracking.xlsx"

def checkForMonday():
    string1day = 'AND JobTrans.ScanDate >= DATEADD(day, -1, GETDATE())'
    string3day = 'AND JobTrans.ScanDate >= DATEADD(day, -3, GETDATE())'
    if datetime.today().weekday() == 0:
        day_string = string3day
#        print ("It's Monday")
    else:
        day_string = string1day
#        print("It's not Monday")
    return day_string

def getData():
    jobslist = {}
    day_string = checkForMonday()        
    conn = pyodbc.connect('Driver={SQL Server};'
                          'Server=SVR-APP\\SQLEXPRESS;'
                          'Database=QssCatiJobTrack;'
                          'Uid=sa;'
                          'Pwd=$ql4Q$$4C@T1;'
                          'Trusted_Connect=yes;')

    cursor = conn.cursor()
    #the query below 
    cursor.execute('SELECT Jobs.JobNum, Parts.CustName, Parts.CustPartNum '
                   'FROM QssCatiJobTrack.dbo.JobTrans, QssCatiJobTrack.dbo.Jobs, QssCatiJobTrack.dbo.Parts '
                   'WHERE Jobs.JobID=JobTrans.JobID '
                   'AND Jobs.PartID = Parts.PartID '
                   'AND JobTrans.ScanFromProcessID=127 '
                   + day_string)


    for row in cursor:
        jobslist[row[0]] = [row[1],row[2]]    
    cursor.close()
    conn.close()    
    return jobslist

def compareCells(sh, jlist):
    for row in sh.iter_rows(min_row = 2, max_row = sh.max_row, max_col = 1):
        for cell in row:
            e = str(cell.value)
            e = e.strip()            
            if e.upper() in jlist or e.lower() in jlist:
                jlist.pop(e.upper(), None)
                jlist.pop(e.lower(), None)    
    return jlist

def countCells(sh):
    count = 0
    for row in sh.iter_rows(min_row = 2, max_row = sh.max_row, max_col =1):
        for cell in row:
            count += 1
    return count    
                    
def writeNTJobs(sh, jlist, today_date):
    write_row = sh.max_row+1
    for e in jlist:
        sh.cell(row=write_row, column=1).value = e
        sh.cell(row=write_row, column=2).value = today_date
        write_row += 1
    jlist = []
    return jlist

def formatMessage(jlist):
    #this returns a string containing dict data in a more human readable format
    output = ""
    
    for k in sorted(jlist):
        job = k
        cust = jlist[k][0]
        p_num = jlist[k][1]
        output = output + 'Job # ' +job + ', Customer: ' + cust + ', Part # ' + p_num + '\n'
    return output

def checkTracked(jlist):
    check_date = datetime.today().date()
    wb = openpyxl.load_workbook(setupPath)
    tsh = wb["Tracked Setups"]
    wipsh = wb["Setups In Progress"]
    ntsh = wb["NonTracked Setups"]    
    jlist = compareCells(tsh, jlist)
    jlist = compareCells(wipsh, jlist)
    jlist = compareCells(ntsh, jlist)

    tracked_count = countCells(tsh)    
    
    if jlist:
        writeNTJobs(ntsh, jlist, check_date)
        non_tracked_count = countCells(ntsh)
        ratio = float(tracked_count) / float(tracked_count + non_tracked_count)
        msg = found_msg + formatMessage(jlist)
        send_email.mailMessage(send_to, subject, msg)
    else:
        non_tracked_count = countCells(ntsh)
        ratio = float(tracked_count) / float(tracked_count + non_tracked_count)
        msg = default_msg
        send_email.mailMessage(send_to, subject, msg)

    wb.save(setupPath)
        
            
job_list = getData()
checkTracked(job_list)
